# Fichier : views/pointage_details.py

from datetime import datetime
import flet as ft
import openpyxl
from t_service_data import read_students_data, delete_pointage_entry
from collections import Counter

# --- FONCTION DE LECTURE DES DONNÉES DE POINTAGE ---
# Cette fonction est modifiée pour être plus robuste et éviter les erreurs.
def read_pointage_data():
    """
    Lit les données de la feuille 'pointage' et les retourne sous forme de liste de dictionnaires.
    Version corrigée et plus robuste.
    """
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        if 'pointage' not in workbook.sheetnames:
            print("Erreur: La feuille 'pointage' n'existe pas.")
            return []
        
        sheet = workbook['pointage']
        
        if sheet.max_row < 2:
            return []
            
        headers = [str(cell.value).strip() for cell in sheet[1]]
        
        data = []
        for row in sheet.iter_rows(min_row=2):
            pointage_dict = {
                headers[i]: cell.value for i, cell in enumerate(row) if i < len(headers)
            }
            # Vérifier si au moins une valeur significative est présente
            if any(pointage_dict.values()):
                data.append(pointage_dict)
        
        return data
        
    except FileNotFoundError:
        print("Erreur: Le fichier 'students.xlsx' n'a pas été trouvé.")
        return []
    except Exception as ex:
        print(f"Erreur lors de la lecture du fichier Excel: {ex}")
        return []

# --- NOUVELLE FONCTION POUR LE FORMATAGE DE LA DATE ---
def format_date_for_display(date_brute):
    if date_brute:
        try:
            date_object = datetime.strptime(str(date_brute).split(' ')[0], '%Y-%m-%d')
            return date_object.strftime('%d-%m-%Y')
        except (ValueError, TypeError) as e:
            print(f"Erreur de format de date: {e} pour la date '{date_brute}'")
            return "Date non valide"
    return "Date non enregistrée"
# ----------------------------------------------------


def pointage_details_view(page, student_id, seance):

    all_pointage_data = read_pointage_data()
    all_students_data = read_students_data()
    
    student_info = next((s for s in all_students_data if str(s.get('id_eleve')).strip() == str(student_id).strip()), None)
    student_name = student_info.get('nom', 'Élève inconnu') if student_info else 'Élève inconnu'
    student_no = student_id.split("_")[-1]
        
    # Filtrer les données de pointage pour cet élève
    student_pointage_history = [
        p for p in all_pointage_data if str(p.get('id_eleve')).strip() == str(student_id).strip()
    ]

    pointage_counts = Counter(p.get('item', 'N/A') for p in student_pointage_history)
    
    list_view = ft.ListView(
        controls=[],
        expand=1,
        spacing=1,
        padding=5,
    )
    
    def handle_delete(e):
        entry_to_delete = e.control.data
        if delete_pointage_entry(
            student_id=entry_to_delete['id_eleve'],
            date_str=entry_to_delete['date'],
            seance=entry_to_delete['seance'],
            item=entry_to_delete['item']
        ):
            snack_delete = ft.SnackBar(ft.Text("Entrée de pointage supprimée avec succès."))
            page.open(snack_delete)
            student_pointage_history.remove(entry_to_delete)
            list_view.controls.clear()
            list_view.controls.extend(create_history_tiles())
            
            pointage_counts.clear()
            pointage_counts.update(p.get('item', 'N/A') for p in student_pointage_history)
            page.views.clear()
            page.views.append(pointage_details_view(page, student_id, seance))
        else:
            snack_delete = ft.SnackBar(ft.Text("Échec de la suppression de l'entrée."))
            page.open(snack_delete)
        page.update()

    def create_history_tiles():
        
        return [
            ft.Card(
                elevation=5,
                content=ft.Container(
                    bgcolor="#C80815",
                    border_radius=8,
                    content=ft.ListTile(
                        title=ft.Text(f"Statut: {p.get('item', 'N/A')}", weight=ft.FontWeight.BOLD, color="white"),
                        subtitle=ft.Text(
                            f"Le {format_date_for_display(p.get('date'))} / {p.get('seance', 'N/A')}", 
                            color="white"
                        ),
                        leading=ft.Icon(
                            ft.Icons.CHECK_CIRCLE,
                            color=ft.Colors.WHITE
                        ),
                        trailing=ft.IconButton(
                            icon=ft.Icons.DELETE_FOREVER_ROUNDED,
                            icon_color=ft.Colors.WHITE,
                            tooltip="Supprimer l'entrée",
                            on_click=handle_delete,
                            data=p
                        ),
                    ),
                    padding=5
                ),
                margin=ft.margin.symmetric(vertical=5, horizontal=10)
            )
            for p in student_pointage_history
        ]

    list_view.controls.extend(create_history_tiles())

    bilan_rows = []
    for item, count in pointage_counts.items():
        bilan_rows.append(
            ft.Row(
                controls=[
                    ft.Text(f"{item}:", size=16, weight=ft.FontWeight.BOLD),
                    ft.Text(f"{count} fois", size=16),
                ],
                alignment=ft.MainAxisAlignment.START,
                spacing=10
            )
        )

    bilan_section = ft.Container(
        content=ft.Column(
            controls=[
                ft.Divider(height=10, color=ft.Colors.GREY_300),
                ft.Text("Bilan d'assiduité", size=18, weight=ft.FontWeight.W_700, text_align=ft.TextAlign.CENTER),
                ft.Divider(height=10, color=ft.Colors.GREY_300),
                *bilan_rows,
            ],
            spacing=10
        ),
        padding=ft.padding.all(20),
    )

    return ft.View(
        f"/pointage_details/{student_id}/{seance}",
        [
            ft.AppBar(
                title=ft.Text(f"Historique de pointage", color=ft.Colors.WHITE),
                bgcolor="#c80815",
                center_title=True,
                leading=ft.IconButton(
                    icon=ft.Icons.ARROW_BACK,
                    icon_color="white",
                    on_click=lambda _: page.go(f"/pointage/{student_info.get('classe', '')}/{seance}")
                )
            ),
            ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Container(
                            content=ft.Text(f"{student_name}", color=ft.Colors.BLACK, size=22),
                            alignment=ft.alignment.center
                        ),
                        ft.Container(
                            content=ft.Text(f"N°: {student_no}", color=ft.Colors.BLACK, size= 18, weight=ft.FontWeight.W_700),
                            alignment=ft.alignment.center
                        ),
                    ]
                )
            ),
            list_view,
            bilan_section,
            ft.Container(
                content=ft.Text(
                    "Aucun historique de pointage pour cet élève.",
                    color=ft.Colors.GREY_500,
                    text_align=ft.TextAlign.CENTER
                ),
                alignment=ft.alignment.center,
                visible=len(student_pointage_history) == 0,
                padding=ft.padding.all(20)
            )
        ],
        bgcolor="#F8F8FF"
    )