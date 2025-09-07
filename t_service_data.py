# Fichier : t_service_data.py

import openpyxl
from datetime import datetime, date
import logging


def read_t_service_data():
    """
    Lit les données de la feuille 't_service' et les retourne sous forme de liste de dictionnaires.
    Nettoie les en-têtes pour éviter les erreurs de clé.
    """
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        if 't_service' not in workbook.sheetnames:
            print("Erreur: La feuille 't_service' n'existe pas.")
            return []
        
        sheet = workbook['t_service']
        
        if sheet.max_row < 2:
            return []
            
        # Nettoyer et normaliser les en-têtes
        headers = [str(cell.value).strip().lower() for cell in sheet[1]]
        data = []
        for row in sheet.iter_rows(min_row=2):
            service_dict = {}
            if row[0].value is None:
                continue
                
            for header, cell in zip(headers, row):
                service_dict[header] = cell.value
                
            data.append(service_dict)
                
        return data
        
    except FileNotFoundError:
        print("Erreur: Le fichier 'students.xlsx' n'a pas été trouvé.")
        return []
    except Exception as ex:
        print(f"Erreur lors de la lecture de la feuille 't_service': {ex}")
        return []
    
# Fonction pour lire les données de la feuille 'students' et les retourne sous forme de liste de dictionnaires.

def read_students_data():
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        sheet = workbook['student']
        
        # Nous allons lire les en-têtes sans normalisation
        headers = [str(cell.value).strip() for cell in sheet[1]]

        data = []
        for row in sheet.iter_rows(min_row=2):
            student_dict = {}
            if row[0].value is None:
                continue
                
            for header, cell in zip(headers, row):
                value = cell.value
                # Normalisation désactivée pour le moment
                student_dict[header] = value
                
            data.append(student_dict)
            
        return data
        
    except Exception as ex:
        print(f"Erreur lors de la lecture du fichier Excel: {ex}")
    return []

#-------------Fin read_students_data()---------------------------
     
def delete_pointage_entry(student_id, date_str, seance, item):

    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        if 'pointage' not in workbook.sheetnames:
            print("Erreur: La feuille 'pointage' n'existe pas.")
            return False
            
        sheet = workbook['pointage']
        
        # Trouver la ligne à supprimer
        row_to_delete = -1
        for row_index, row in enumerate(sheet.iter_rows(min_row=2)):
            # On vérifie si la ligne correspond à l'entrée
            # On utilise str() et strip() pour s'assurer que les comparaisons sont exactes
            if (str(row[0].value).strip().lower() == str(student_id).strip().lower() and
                str(row[1].value).strip().lower() == str(date_str).strip().lower() and
                str(row[2].value).strip().lower() == str(seance).strip().lower() and
                str(row[3].value).strip().lower() == str(item).strip().lower()):
                
                row_to_delete = row_index + 2  # +2 car le décompte commence à 0 et on ignore la première ligne d'en-tête
                break
        
        if row_to_delete != -1:
            sheet.delete_rows(row_to_delete)
            workbook.save("students.xlsx")
            return True
        else:
            print("Erreur: L'entrée de pointage n'a pas été trouvée.")
            return False
            
    except Exception as ex:
        print(f"Erreur lors de la suppression de l'entrée: {ex}")
        return False

def read_pointage_data():
    """
    Lit les données de pointage depuis le fichier Excel et les retourne
    sous forme de dictionnaire imbriqué.
    """
    pointage_data = {}
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        
        if 'pointage' not in workbook.sheetnames:
            # Si la feuille n'existe pas, la créer immédiatement
            logging.info("Feuille 'pointage' non trouvée. Création de la feuille...")
            workbook.create_sheet('pointage')
            sheet = workbook['pointage']
            sheet.append(['id_eleve', 'date', 'seance', 'item', 'classe'])
            workbook.save("students.xlsx")
            return {} # Retourne un dictionnaire vide après la création
        
        logging.info("Feuille 'pointage' trouvée. Lecture des données...")
        sheet = workbook['pointage']

        if sheet.max_row > 1:
            headers = [cell.value for cell in sheet[1]]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = dict(zip(headers, row))
                
                student_id = row_dict.get('id_eleve')
                
                attendance_date = row_dict.get('date')
                if isinstance(attendance_date, datetime):
                    attendance_date = attendance_date.date().isoformat()
                elif not isinstance(attendance_date, str):
                    continue

                classe = str(row_dict.get('classe')).strip().lower() if row_dict.get('classe') is not None else None
                seance = str(row_dict.get('seance')).strip() if row_dict.get('seance') is not None else None
                item = str(row_dict.get('item')).strip() if row_dict.get('item') is not None else None

                if attendance_date and student_id and seance and item and classe:
                    if attendance_date not in pointage_data:
                        pointage_data[attendance_date] = {}
                    if classe not in pointage_data[attendance_date]:
                        pointage_data[attendance_date][classe] = {}
                    if seance not in pointage_data[attendance_date][classe]:
                        pointage_data[attendance_date][classe][seance] = {}

                    pointage_data[attendance_date][classe][seance][student_id] = {'statut': item}
        logging.info(f"Données de pointage chargées. Nombre de dates trouvées : {len(pointage_data)}")
        return pointage_data
    except Exception as e:
        logging.error(f"Erreur inattendue lors de la lecture du pointage : {e}")
        return {}

#Fonction qui lit les absents depuis le fichier Excel    
def get_absent_students_from_excel(classe, seance):
    """Lit le fichier Excel et retourne un dictionnaire des élèves absents pour la date,
    la classe et la séance données."""
    absent_students_data = {}
    current_date_str = date.today().isoformat()
    
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        sheet = workbook['pointage']
        
        headers = [cell.value for cell in sheet[1]]
        
        for row in sheet.iter_rows(min_row=2):
            row_data = {headers[i]: cell.value for i, cell in enumerate(row)}
            
            # Vérification des conditions : date, classe, séance et statut
            if (str(row_data.get('date')).strip() == current_date_str and
                str(row_data.get('classe')).strip() == str(classe).strip() and
                str(row_data.get('seance')).strip() == str(seance).strip() and
                row_data.get('item') in ["Absent", "Sanctionné"]):
                
                student_id = str(row_data.get('id_eleve')).strip()
                item = row_data.get('item')
                absent_students_data[student_id] = {'item': item}
                
    except Exception as ex:
        print(f"Erreur lors de la lecture des données de pointage: {ex}")
    
    return absent_students_data

#Fonction de vérification de doublon
def check_student_id_exists(student_id):
    """
    Vérifie si un élève avec cet ID existe déjà.
    Cette version est la plus robuste car elle nettoie les IDs avant de les comparer.
    """
    all_students = read_students_data()
    
    # Nettoyer l'ID saisi par l'utilisateur
    cleaned_input_id = str(student_id).strip().replace(" ", "")

    for student in all_students:
        existing_id = student.get('id_eleve', '')
        # Nettoyer l'ID existant dans le fichier
        cleaned_existing_id = str(existing_id).strip().replace(" ", "")
        
        # La comparaison est maintenant une égalité stricte entre deux IDs nettoyés
        if cleaned_existing_id == cleaned_input_id:
            return True  # L'ID existe déjà
            
    return False  # L'ID n'existe pas

#
def check_student_number_in_class(classe, student_number):
    """
    Vérifie si un élève avec le même numéro existe déjà dans une classe donnée.
    Cette version est plus robuste et évite les faux positifs.
    """
    all_students = read_students_data()
    for student in all_students:
        student_id = str(student.get('id_eleve', '')).strip()
        
        # On s'assure que l'ID a le bon format et on extrait le numéro
        if '_' in student_id:
            existing_classe, existing_number = student_id.split('_', 1)
            
            # On compare la classe et le numéro exacts
            if existing_classe == classe and existing_number == student_number:
                return True # Un élève avec le même numéro dans cette classe a été trouvé
                
    return False # Aucun doublon trouvé