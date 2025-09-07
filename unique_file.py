# Nouveau fichier Python unique
import flet as ft
import openpyxl
import pandas as pd
from views.menu import menu_view
from views.home import home_view
from views.details import details_view
from views.assiduite import assiduite_view
from views.edit_student_view import edit_student_view
from views.add_student_view import add_student_view

# ... read_students_data
# fonction pour lire le fichier excel et le convertir en dictionnaires
def read_students_data():
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        sheet = workbook.active
        
        data = []
        headers = [cell.value for cell in sheet[1]]
        
        # On lit les données à partir de la deuxième ligne
        for row in sheet.iter_rows(min_row=2):
            student_dict = {}
            for header, cell in zip(headers, row):
                student_dict[header] = cell.value
            data.append(student_dict)
            
        return data
    except FileNotFoundError:
        print("Erreur: Le fichier 'students.xlsx' n'a pas été trouvé.")
        return []
    except Exception as ex:
        print(f"Erreur lors de la lecture du fichier Excel: {ex}")
        return []

# ... get_student_by_id
def get_student_by_id(page, student_id):
    # Elle utilise la liste de la page, qui est toujours à jour
    for student in page.students_data:
        if str(student['id']) == student_id:
            return student
    return None
# ... delete_and_refresh
def delete_and_refresh(page, student_id_to_delete, classe):
    try:
        # Votre code pour supprimer l'élève du fichier Excel
        workbook = openpyxl.load_workbook("students.xlsx")
        sheet = workbook.active
        
        row_to_delete_index = -1
        for row_index, row in enumerate(sheet.iter_rows()):
            if row[0].value == student_id_to_delete:
                row_to_delete_index = row_index + 1
                break
        
        if row_to_delete_index != -1:
            sheet.delete_rows(row_to_delete_index)
            workbook.save("students.xlsx")
        
        # Mettre à jour la liste en mémoire (l'état de la page)
        page.students_data = [s for s in page.students_data if s['id'] != student_id_to_delete]
        
        # Afficher un message de succès
        page.snack_bar = ft.SnackBar(ft.Text("Élève supprimé avec succès !"))
        page.snack_bar.open = True

        # Forcer la navigation vers la même page pour la rafraîchir
        page.go(f"/home/{classe}")

    except Exception as ex:
        page.snack_bar = ft.SnackBar(ft.Text(f"Erreur lors de la suppression: {ex}"))
        page.snack_bar.open = True
        page.update()

# ... create_confirm_dialog
def create_confirm_dialog(page, student_id, student_name, classe):
    
    def close_dialog(e):
        e.page.dialog.open = False
        e.page.update()

    def confirm_delete(e):
        close_dialog(e)
        # On appelle la nouvelle fonction qui gère tout
        delete_and_refresh(e.page, student_id, classe)
    
    return ft.AlertDialog(
        modal=True,
        title=ft.Text("Confirmer la suppression"),
        content=ft.Text(f"Êtes-vous sûr de vouloir supprimer l'élève {student_name} ? Cette action est irréversible."),
        actions=[
            ft.TextButton("Annuler", on_click=close_dialog),
            ft.TextButton("Confirmer", on_click=confirm_delete),
        ],
        actions_alignment=ft.MainAxisAlignment.END,
    )
# ... create_student_list_tile
def create_student_list_tile(page, student, classe):
    
    # Définir le comportement du bouton DELETE
    def on_delete_click(e):
        dialog = create_confirm_dialog(page, student['id'], student['nom'], classe)
        page.dialog = dialog
        page.open(dialog)
        page.update()

    # Bouton DELETE qui appelle la fonction de suppression
    delete_button = ft.IconButton(
        icon=ft.Icons.DELETE,
        icon_color=ft.Colors.RED_500,
        data=student['id'],  # On stocke l'ID de l'étudiant dans le bouton
        on_click=on_delete_click,
        
    )
    
    # Bouton EDIT qui va à la page de modification
    edit_button = ft.IconButton(
        icon=ft.Icons.EDIT,
        icon_color=ft.Colors.WHITE,
        on_click=lambda _, student_id=student['id']: page.go(f"/student_edit/{student_id}")
    )
    avatar_url = "assets/avatar_G.png" if student['genre'] == "G" else "assets/avatar_F.png"

    return ft.Container(
        content=ft.Row(
            controls=[
                #Display_data_Container
                ft.Container(
                    content=ft.ListTile(
                    title=ft.Text(f"{student['id'].split('_')[-1]} : {student['nom']}", color="white"),
                    subtitle=ft.Text(f"Genre: {student['genre']} ", color="white"),
                    leading=ft.CircleAvatar(
                        content=ft.Image(
                            src=avatar_url,
                            fit=ft.ImageFit.COVER
                        ),
                        radius=25
                    ),
        
                ),
                    expand=True,

            ),
            # Edit_Container
                ft.Container(
                    content=ft.Column(
                        controls=[
                            edit_button,
                            delete_button,
                        ]
                    ),
                   ),
            ]
        ),
        alignment=ft.alignment.center,
        padding=3,
        bgcolor="#002e63",
        border_radius=4,
        on_click=lambda _, student_id=student['id']: page.go(f"/student/{student_id}"),
    )
# Fonction pour créer la barre de titre de la vue
def create_title_bar(page, title):
    """
    Crée le AppBar pour la vue.
    """
    return ft.AppBar(
        title=ft.Text("Liste des élèves de la classe" , color="white"),
        leading=ft.IconButton(
            icon=ft.Icons.ARROW_BACK,
            on_click=lambda _: page.go("/"),
            icon_color="white"
        ),
        center_title=True
    )
# ... home_view
def home_view(page, classe):
    """
    Retourne la vue complète pour la page d'accueil de la classe.
    """
    students_list_tiles = [
        # Appel de la fonction pour chaque étudiant
        create_student_list_tile(page, student, classe)
        for student in page.students_data if student['classe'] == classe
    ]

    return ft.View(
        f"/home/{classe}",
        [
            create_title_bar(page, "Liste des élèves de la classe"),
            # Les autres contrôles sont ajoutés directement
            ft.Container(
                content=ft.Stack(
                    controls=[
                        ft.Image(
                            src="assets/image_fond1.png",
                            fit=ft.ImageFit.FILL,
                            height=60,
                            width=page.window.width
                        ),
                        ft.Container(
                            content=ft.Text(f"---{classe}---", size=22, weight=ft.FontWeight.W_500),
                            alignment=ft.alignment.center,
                        )
                    ],
                    height=60,
                ),
                alignment=ft.alignment.center,
            ),
            ft.Column(
                controls=[
                    ft.ListView(
                        controls=students_list_tiles,
                        spacing=2,
                        expand=True,
                    )
                ],
                expand=True,
            ),
            
        ],
        floating_action_button=ft.FloatingActionButton(
            icon=ft.Icons.ADD,
            on_click=lambda _,:page.go(f"/add_student/{classe}"),
            tooltip="Ajouter un nouvel élève",
            
        ),
       floating_action_button_location=ft.FloatingActionButtonLocation.CENTER_DOCKED 
    )
# fonction main
def main(page: ft.Page):
    page.title='Navigation avec passage de paramètres...'
    page.window.width = 400
    page.vertical_alignment = ft.MainAxisAlignment.CENTER
    page.horizontal_alignment = ft.CrossAxisAlignment.CENTER
    page.scroll = "auto"
    
    appbar_theme = ft.AppBarTheme(
        toolbar_height = 400,
        bgcolor="#002e63"
    )
    page.theme = ft.Theme(appbar_theme=appbar_theme)

    # CHARGEMENT DES DONNÉES UNE SEULE FOIS, DIRECTEMENT SUR L'OBJET PAGE
    page.students_data = read_students_data()

    def route_change(e):
        page.views.clear()
        if page.route == "/":
            page.views.append(menu_view(page))
        elif page.route.startswith("/add_student/"):
            classe = page.route.split("/")[-1]
            page.views.append(add_student_view(page, classe))
        elif page.route.startswith("/home/"):
            classe = page.route.split("/")[-1]
            page.views.append(home_view(page, classe))
        elif page.route.startswith("/student/"):
            student_id = page.route.split("/")[-1]
            # ON PASSE L'OBJET PAGE À LA FONCTION
            student = get_student_by_id(page, student_id)
            if student:
                page.views.append(details_view(page, student, student_id))
            else:
                page.views.append(ft.View("/404", [ft.AppBar(title=ft.Text("Erreur")), ft.Text("Étudiant non trouvé")]))
            
        elif page.route.startswith("/assiduité/"):
            student_id = page.route.split("/")[-1]
            # ON PASSE L'OBJET PAGE À LA FONCTION
            student = get_student_by_id(page, student_id)
            if student:
                page.views.append(assiduite_view(page, student, student_id))
            else:
                page.views.append(ft.View("/404", [ft.AppBar(title=ft.Text("Erreur")), ft.Text("Étudiant non trouvé")]))
        elif page.route.startswith("/student_edit/"):
            student_id = page.route.split("/")[-1]
            page.views.append(edit_student_view(page, student_id)) 
        
        page.update()

    def view_pop(e):
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    page.on_route_change = route_change
    page.on_view_pop = view_pop
    page.go(page.route)


ft.app(target=main, assets_dir="assets")