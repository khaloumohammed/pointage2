# Fichier : views/pointage.py

import flet as ft
import openpyxl
from datetime import date, datetime
from t_service_data import read_students_data, get_absent_students_from_excel

# La fonction record_pointage doit recevoir la classe pour l'enregistrer dans Excel
def record_pointage(student_id, item, seance, classe):
    try:
        workbook = openpyxl.load_workbook("students.xlsx")
        if 'pointage' not in workbook.sheetnames:
            workbook.create_sheet('pointage')
            sheet = workbook['pointage']
            sheet.append(['id_eleve', 'date', 'seance', 'item', 'classe'])
        else:
            sheet = workbook['pointage']
            
        new_row = [student_id, date.today().isoformat(), seance, item, str(classe).strip()]             
        sheet.append(new_row)
        
        workbook.save("students.xlsx")
        return True
        
    except Exception as ex:
        print(f"Erreur lors de l'enregistrement du pointage: {ex}")
        return False

# Fonction pour afficher la boîte de dialogue des absents
def show_absent_students_dialog(page, current_classe, current_seance):
    current_date = date.today().isoformat()
    # Vous pouvez formater la date pour une meilleure lisibilité
    formatted_date = date.today().strftime('%d-%m-%Y')
    # Appel de la nouvelle fonction pour lire les données depuis le fichier Excel
    session_absences = get_absent_students_from_excel(current_classe, current_seance)
    
    absent_list_controls = []
    
    for student_id, data in session_absences.items():
        # La logique de recherche est la même, mais les données viennent du fichier
        cleaned_absent_id = student_id
        
        student = next(
            (s for s in page.students_data if s.get('id_eleve', '').strip() == cleaned_absent_id), 
            None
        )
        
        if student:
            student_name = student.get('nom', 'Nom non trouvé')
            id_display = student_id.split("_")[-1]
            absent_list_controls.append(
                ft.Text(f"N° {id_display} : {student_name}", size=16) 
            )
        else:
            absent_list_controls.append(
                ft.Text(f"N° {student_id.split('_')[-1]} : Nom non trouvé", size=16)
            )

    if not absent_list_controls:
        absent_list_controls.append(ft.Text("Aucun élève absent pour l'instant."))



    dialog = ft.AlertDialog(
        modal=True,
        title=ft.Text(
            f" {formatted_date} \n {current_classe} - {current_seance} \n {len(session_absences)} absent(s)",
            size= 16,
            weight=ft.FontWeight.W_600,
            text_align=ft.TextAlign.CENTER,
            ),
        
        content=ft.Column(
            absent_list_controls,
            height=200,
            scroll="always"
        ),
        actions=[
            ft.TextButton(
                "Fermer",
                style= ft.ButtonStyle(color="black")  ,
                on_click=lambda e: close_dialog(page, dialog)),
        ],
        actions_alignment="center",
        scrollable= True,
        
    )
    
    page.dialog = dialog
    page.open(dialog)
    page.update()

def close_dialog(page, dialog):
    page.close(dialog)
    page.update()

def pointage_view(page, classe, seance):
    formatted_classe = str(classe).strip()
    
    all_students_data = read_students_data()
    students_in_class = [s for s in all_students_data if str(s.get('classe')).strip() == formatted_classe]
    #---------------------
    def create_student_pointage_card(student):
        def handle_click(e):
            item = e.control.text
            # Utiliser .strip() pour éviter les espaces inutiles
            student_id = str(student.get('id_eleve', '')).strip()

            # On passe la classe comme argument à la fonction de sauvegarde
            if record_pointage(student_id, item, seance, classe):
                current_date = date.today().isoformat()
                formatted_classe = str(classe).strip()
                
                # S'assurer que la structure existe
                if current_date not in page.absences:
                    page.absences[current_date] = {}
                if formatted_classe not in page.absences[current_date]:
                    page.absences[current_date][formatted_classe] = {}
                if seance not in page.absences[current_date][formatted_classe]:
                    page.absences[current_date][formatted_classe][seance] = {}
                if student_id not in page.absences[current_date][formatted_classe][seance]:
                    page.absences[current_date][formatted_classe][seance][student_id] = {}

                # Mise à jour de l'état en mémoire
                page.absences[current_date][formatted_classe][seance][student_id]['item'] = item
                
                snack_pointage = ft.SnackBar(ft.Text(f"Pointage '{item}' enregistré pour {student.get('nom')}"))
                page.open(snack_pointage)
            else:
                snack_pointage = ft.SnackBar(ft.Text(f"Échec de l'enregistrement du pointage."))
                page.open(snack_pointage)
            page.update()
        
        student_id = student.get('id_eleve', 'N/A')
        id_display = student_id.split("_")[-1]
        student_name = student.get('nom', 'Nom non trouvé')
        student_genre = student.get('genre', 'g')
        avatar_url = "assets/avatars/avatar_G.png" if student_genre == "G" else "assets/avatars/avatar_F.png"
        
        pointage_buttons = ft.Container(
            expand=True,
            padding=2,
            alignment=ft.alignment.center,
            content=ft.Column(
                alignment=ft.MainAxisAlignment.CENTER,
                controls=[
                    ft.Row(
                        alignment=ft.MainAxisAlignment.CENTER,
                        controls=[
                            ft.TextButton(text="Présent", style=ft.ButtonStyle(color="white", bgcolor="#C80815"), on_click=handle_click),
                            ft.TextButton(text="Absent", style=ft.ButtonStyle(color="white", bgcolor="#C80815"), on_click=handle_click),
                            ft.TextButton(text="En retard", style=ft.ButtonStyle(color="white", bgcolor="#C80815"), on_click=handle_click),
                            ft.TextButton(text="En civil", style=ft.ButtonStyle(color="white", bgcolor="#C80815"), on_click=handle_click),
                            ft.TextButton(text="Malade", style=ft.ButtonStyle(color="white", bgcolor="#C80815"), on_click=handle_click),
                            ft.TextButton(text="Sanctionné", style=ft.ButtonStyle(color="white", bgcolor="#C80815"), on_click=handle_click),
                            ft.TextButton(text="Gratifié", style=ft.ButtonStyle(color="white", bgcolor="#C80815"), on_click=handle_click),
                        ],
                        wrap=True,
                        spacing=10,
                    ),
                ]
            )
        )
        
        return ft.Card(
            expand=True,
            elevation=8,
            content=ft.Container(
                bgcolor="#C80815",
                border_radius=8,
                width=page.window.width,
                padding=5,
                expand=True,
                content=ft.Column(
                    controls=[
                        ft.Row(
                            alignment=ft.MainAxisAlignment.SPACE_BETWEEN,
                            controls=[
                                ft.CircleAvatar(
                                    content=ft.Image(
                                        src=avatar_url,
                                        fit=ft.ImageFit.COVER
                                    ),
                                    radius=25
                                ),
                                ft.Column(
                                    controls=[
                                        ft.Text(f"{student_name}", weight=ft.FontWeight.BOLD, size=22, color="white"),
                                        ft.Text(f"N°: {id_display}", size=18, color=ft.Colors.WHITE),
                                    ],
                                    spacing=5,
                                    alignment=ft.MainAxisAlignment.CENTER,
                                    horizontal_alignment=ft.CrossAxisAlignment.CENTER
                                ),
                                ft.TextButton(
                                    text="Historique",
                                    style=ft.ButtonStyle(color="white", bgcolor="#C80815"),
                                    on_click=lambda _: page.go(f"/pointage_details/{student_id}/{seance}"),
                                ),
                            ],
                            spacing=10,
                        ),
                        ft.Divider(height=2, color=ft.Colors.WHITE),
                        pointage_buttons,
                    ]
                )
            ),
            margin=ft.margin.symmetric(vertical=10, horizontal=15),
        )

    pointage_cards = [
        create_student_pointage_card(student) for student in students_in_class
    ]
    
    return ft.View(
        f"/pointage/{classe}/{seance}",
        [
            ft.AppBar(
                title=ft.Text("Pointage", color=ft.Colors.WHITE),
                bgcolor="#c80815",
                center_title=True,
                leading=ft.IconButton(
                    icon=ft.Icons.ARROW_BACK,
                    icon_color="white",
                    on_click=lambda _: page.go("/")
                ),
                actions=[
                    # L'action de l'AppBar qui appelle la boîte de dialogue
                    ft.TextButton(
                        content=ft.Text("Les absents", size=16, color=ft.Colors.WHITE),
                        on_click=lambda _: show_absent_students_dialog(page, classe, seance),
                    ),
                ]
            ),
            ft.Container(
                content=ft.Column(
                    controls=[
                        ft.Container(
                            content=ft.Text(f"{classe}", color=ft.Colors.BLACK, size=22),
                            alignment=ft.alignment.center
                        ),
                        ft.Container(
                            content=ft.Text(f"{seance}", color=ft.Colors.BLACK, size=18, weight=ft.FontWeight.W_700),
                            alignment=ft.alignment.center
                        ),
                    ]
                )
            ),
            ft.ListView(
                controls=pointage_cards,
                expand=True,
                spacing=1,
                padding=5,
            ),
            ft.Container(
                content=ft.Text(
                    f"Aucun élève trouvé pour la classe '{classe}'.",
                    color=ft.Colors.GREY_500,
                    text_align=ft.TextAlign.CENTER
                ),
                alignment=ft.alignment.center,
                visible=len(students_in_class) == 0,
                padding=ft.padding.all(5)
            )
        ],
        spacing=2,
        bgcolor="#F8F8FF",
        padding=5
    )
